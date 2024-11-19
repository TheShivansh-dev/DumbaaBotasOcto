# Token and Bot Username
#TOKEN: Final = '7652253001:AAEipGC5Fb0Y04NgbCICb6N1Tm6HcJG4tpA'
#BOT_USERNAME: Final = '@Dumbaa_bot'
#TOKEN: Final = '7007935023:AAENkGaklw6LMJA_sfhVZhnoAgIjW4lDTBc'
#BOT_USERNAME: Final = '@Grovieee_bot'
#ALLOWED_GROUP_IDS = [-1001817635995, -1002114430690]


import os
import random
import re
import difflib
from typing import Final
import telegram
from telegram import Update, InputFile, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, CallbackQueryHandler, filters, ContextTypes
import openpyxl

EXCEL_FILE = 'user_scores.xlsx'
TOKEN: Final = '7652253001:AAEipGC5Fb0Y04NgbCICb6N1Tm6HcJG4tpA'
BOT_USERNAME: Final = '@Dumbaa_bot'
ALLOWED_GROUP_IDS = [-1001817635995, -1002114430690]
# Dictionary to keep track of ongoing games
octo_game_state = {}

 

# Helper to escape MarkdownV2 characters
def escape_markdown_v2(text: str) -> str:
    return re.sub(r'([_\*\[\]\(\)~`>#+\-=|{}.!])', r'\\\1', text)

async def show_all_results(update: Update, context: ContextTypes.DEFAULT_TYPE):
    scores = load_scores()  # Load scores from the Excel file

    if not scores:
        try:
            await update.message.reply_text("No scores found")
        except telegram.error.BadRequest:
            await update.message.chat.send_message("No scores found")
        return

    # Build the message to display all users
    message = "*All Users and Scores:*\n"
    for user_id, username, score in scores:
        formatted_score = f"{score:.2f}"
        message += f"ID: {user_id}, Username: @{escape_markdown(str(username))}, Score: {escape_markdown(str(formatted_score))}points\n"
    try:
        await update.message.reply_text(message, parse_mode='MarkdownV2')
    except telegram.error.BadRequest:
        await update.message.chat.send_message(message, parse_mode='MarkdownV2')

def update_user_score(user_id: int, username: str, score: float):
    # If the file does not exist, create it with headers

    if not os.path.exists(EXCEL_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Scores'
        # Create the header row
        sheet.append(['Idnumber', 'Username', 'Score'])
        workbook.save(EXCEL_FILE)

    # Load the existing Excel file
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active

    # Check if the user already exists by checking the user ID
    user_found = False
    for row in range(2, sheet.max_row + 1):  # Start from row 2 to skip the header
        if sheet.cell(row=row, column=1).value == user_id:  # The user ID is in the first column (Idnumber)
            # Update the score for this user
            current_score = sheet.cell(row=row, column=3).value
            new_score = current_score + score if current_score is not None else score
            sheet.cell(row=row, column=3, value=new_score)  # Update the score in the 3rd column
            user_found = True
            break

    if not user_found:
        # If user not found, add a new row with the user ID, username, and score
        sheet.append([user_id, username, score])

    # Save the updated Excel file
    workbook.save(EXCEL_FILE)
    workbook.close()

# Load all user scores from the Excel file
def load_scores():
    if not os.path.exists(EXCEL_FILE):
        return []

    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active

    scores = []
    for row in range(2, sheet.max_row + 1):  # Start from row 2 to skip the header
        user_id = sheet.cell(row=row, column=1).value
        username = sheet.cell(row=row, column=2).value
        score = sheet.cell(row=row, column=3).value

        if user_id and username and score is not None:
            scores.append((user_id, username, score))

    workbook.close()
    return scores

# Command to show the top 10 users
async def select_top_10_users(update: Update, context: ContextTypes.DEFAULT_TYPE):
    scores = load_scores()
    if not scores:
        try:
            await update.message.reply_text("No scores found")
        except telegram.error.BadRequest:
            await update.message.chat.send_message("No scores found")
        return
    # Sort by score in descending order
    scores.sort(key=lambda x: x[2], reverse=True)
    # Get the top 10 users
    top_10 = scores[:10]

    # Build the message to display top users
    message = "*Top 10 Dumbaa:*\n"
    for idx, (user_id, username, score) in enumerate(top_10, 1):
        formatted_score = f"{score:.2f}"
        message += f"{idx}: @{escape_markdown(str(username))} : {escape_markdown(str(formatted_score))} points\n"
    try:
        await update.message.reply_text(message, parse_mode='MarkdownV2')
    except telegram.error.BadRequest:
        await update.message.chat.send_message(message, parse_mode='MarkdownV2')

# Command to show the user's rank and score
async def my_rank(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    username = update.message.from_user.username or update.message.from_user.first_name

    scores = load_scores()

    if not scores:
        try:
            await update.message.reply_text("No scores found")
        except telegram.error.BadRequest:
            await update.message.chat.send_message("No scores found")
        return

    # Sort by score in descending order
    scores.sort(key=lambda x: x[2], reverse=True)

    # Find user's rank
    user_rank = None
    for rank, (u_id, u_name, score) in enumerate(scores, 1):
        if u_id == user_id:
            user_rank = (rank, score)
            break

    if user_rank:
        rank, score = user_rank
        formatted_score = f"{score:.2f}"
        try:
            await update.message.reply_text(f"Your rank: {rank}\nYour score: {formatted_score}")
        except telegram.error.BadRequest:
            await update.message.chat.send_message(f"Your rank: {rank}\nYour score: {formatted_score}")
    else:
        try:
            await update.message.reply_text("You haven't played the game yet")
        except telegram.error.BadRequest:
            await update.message.chat.send_message("You haven't played the game yet")

# Function to get a random word from the Excel file
def get_random_word_from_excel(file_path: str, used_srno: list):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Collect words and points from the Excel file
        words_data = []
        for row in range(2, sheet.max_row + 1):  # Start from the second row to skip headers
            srno = sheet.cell(row=row, column=1).value  # 'srno' is in the first column
            if srno in used_srno:  # Skip words that have already been used
                continue
            word = sheet.cell(row=row, column=2).value  # Assuming word is in column 2
            point = sheet.cell(row=row, column=3).value  # Assuming points are in column 3

            # Append tuple of srno, word, and points
            words_data.append((srno, word, point))

        # Choose a random word from the list of unused words
        if words_data:
            srno, word, point = random.choice(words_data)

            return srno, word, point
        else:
            return None, None, None

    except FileNotFoundError:
        return None, None, None

async def start_game_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.message.chat.id
    global dumbai
    dumbai = "dumba"
    # Check if the chat_id (group ID) is in the allowed list
    '''
    if chat_id not in ALLOWED_GROUP_IDS:
        try:
            await update.message.reply_text("Due to the free service, you are not allowed to start a game in this group. Play there https://t.me/+yVFKtplWZUA0Yzhl or contact @O000000000O00000000O")
        except telegram.error.BadRequest:
            await update.message.chat.send_message("Due to the free service, you are not allowed to start a game in this group. Play there https://t.me/+yVFKtplWZUA0Yzhl or contact @O000000000O00000000O")
        return
    '''

    if chat_id in octo_game_state:
        try:
            await update.message.reply_text("A game is already running in this group. hit /cancel to end the game")
        except telegram.error.BadRequest:
            await update.message.chat.send_message("A game is already running in this group. hti /cancel to end the game")
        return

    # Present difficulty selection buttons: Easy or Hard
    difficulty_keyboard = [
        [InlineKeyboardButton("Easy 😄", callback_data='difficulty_easy')],
        [InlineKeyboardButton("Hard 😓", callback_data='difficulty_hard')],
    ]
    reply_markup = InlineKeyboardMarkup(difficulty_keyboard)

    try:
        await update.message.reply_text('Select the difficulty:', reply_markup=reply_markup)
    except telegram.error.BadRequest:
        await update.message.chat.send_message('Select the difficulty:', reply_markup=reply_markup)

async def handle_difficulty_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
   
    await query.answer()
    chat_id = query.message.chat.id
    if chat_id in octo_game_state:
        try:
            await query.message.reply_text("A game is already running in this group hit /cancel to end the game")
        except telegram.error.BadRequest:
            await query.message.chat.send_message("A game is already running in this group hit /cancel to end the game")
        return
    # Determine difficulty based on callback data
    global OCTO_EXCEL_FILE
    difficulty_message = ''
    if query.data == 'difficulty_easy':
        OCTO_EXCEL_FILE = 'octowordexcel.xlsx'  # Use the easy file
        difficulty_message = "Easy mode selected"
    elif query.data == 'difficulty_hard':
        OCTO_EXCEL_FILE = 'hardumbaword.xlsx'  # Use the hard file
        difficulty_message = "Hard mode selected"

    # Ask how many words to play with
    keyboard = [
        [InlineKeyboardButton("25 Words", callback_data='25')],
        [InlineKeyboardButton("100 Words", callback_data='100')],
        [InlineKeyboardButton("250 Words", callback_data='250')],
        [InlineKeyboardButton("500 Words", callback_data='500')],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    # Send a confirmation message along with the next prompt
    try:
        await query.message.reply_text(f"{difficulty_message} How many words do you want", reply_markup=reply_markup)
    except telegram.error.BadRequest:
        await query.message.chat.send_message(f"{difficulty_message} How many words do you want", reply_markup=reply_markup)

async def cancel_game(update: Update, context: ContextTypes.DEFAULT_TYPE):

    chat_id = update.message.chat.id

    # Check if there is an ongoing game in this chat
    if chat_id in octo_game_state:
        # Show the game results before canceling
        await show_game_results(update.message, chat_id)

        # Clear the game state for this chat
        del octo_game_state[chat_id]
        try:
            await update.message.reply_text("The game has been canceled You can start a new game with /startdumba")
        except telegram.error.BadRequest:
            await update.message.chat.send_message("The game has been canceled You can start a new game with /startdumba")
    else:
        try:
            await update.message.reply_text("There is no game currently running in this chat")
        except telegram.error.BadRequest:
            await update.message.chat.send_message("There is no game currently running in this chat")

def is_similar_word_in_message(user_text: str, word: str, threshold: float = 0.7) -> bool:
    """
    Check if the user's text contains the word with a similarity above the given threshold.
    First, attempt to match the word exactly (ignoring case and spaces). If not an exact match,
    check for similarity above the threshold.
    """
    print("Enter in Similar word",user_text,word)
    # Convert both user text and the word to lowercase and strip leading/trailing spaces
    user_text = user_text.lower().strip()
    word = word.lower().strip()

    # Check for an exact match
    if user_text == word:
        return True

async def process_game_round(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Process the user's guess during the octo game round
    """
    if update.message is None:
        return
    print(("Enter in PRocess Round"))
    chat_id = update.message.chat.id
    message = update.message.text.strip()
    user_id = update.message.from_user.id
    username = update.message.from_user.username or update.message.from_user.first_name

    if chat_id not in octo_game_state:
        return

    game_state = octo_game_state[chat_id]
    current_word = game_state['current_word']
    total_rounds = game_state['total_rounds']

    # Initialize players if not already done
    if 'players' not in game_state:
        game_state['players'] = {}

    # Initialize the player's data if not present
    if user_id not in game_state['players']:
        game_state['players'][user_id] = {
            'username': username,
            'current_game_score': 0
        }

    # Check if the user's message contains the word
    if is_similar_word_in_message(message, current_word):
        # User guessed correctly, award points
        points = game_state['current_points']
        game_state['players'][user_id]['current_game_score'] += points  # Update current game score
        update_user_score(user_id, username, points)
        formatted_score = f"{points:.2f}"


        # Proceed to the next round
        game_state['current_round'] += 1

        # If the game is still ongoing, provide the next word
        if game_state['current_round'] <= total_rounds:

            if dumbai =="dumba":
                print(("Enter in PRocess Round 1",dumbai))
                next_srno, next_word, next_points = get_random_word_from_excel(OCTO_EXCEL_FILE, game_state.get('used_srno', []))

                if next_word:
                    game_state.setdefault('used_srno', []).append(next_srno)
                    game_state['current_word'] = next_word
                    game_state['current_points'] = next_points

                    # Generate the scrambled word and the masked word
                    scrambled_word = ' '.join(random.sample(next_word, len(next_word)))
                    masked_word = mask_word(next_word)  # Call to the mask_word function

                    pass_button = InlineKeyboardButton(text="Pass 🐛", callback_data="pass")
                    keyboard = InlineKeyboardMarkup([[pass_button]])

                    # Send the response with both the scrambled word and masked word
                    try:
                        await update.message.reply_text(
                            f"@{username} earned {formatted_score} points for: {current_word}\n\n"
                            f"👻 Round: {game_state['current_round']}/{total_rounds}.\n"
                            f"🎖️ Points: {next_points}\n"
                            f"📚 Letters: {scrambled_word}\n"
                            f"🎲 Guess: {masked_word}\n",
                            reply_markup=keyboard
                        )
                    except telegram.error.BadRequest:
                        # If replying fails, send a normal message
                        await update.message.chat.send_message(
                            f"Correct! @{username} earned {formatted_score} points for:→ {current_word}\n\n"
                            f"👻 Round: {game_state['current_round']}/{total_rounds}.\n"
                            f"🎖️ Points: {next_points}\n"
                            f"📚 Letters: {scrambled_word}\n"
                            f"🎲 Guess: {masked_word}\n",
                            reply_markup=keyboard
                        )
                else:
                    # If no more words are available, end the game
                    try:
                        await update.message.reply_text("No more words available. The game is over.")
                    except telegram.error.BadRequest:
                        await update.message.chat.send_message("No more words available. The game is over.")
                    await show_game_results(update.message, chat_id)  # Show results
                    del octo_game_state[chat_id]
            elif dumbai =="dumbii":
                print(("Enter in PRocess Round 2",dumbai))
                next_srno,next_definition, next_answer, next_point = dumbi_get_random_word_from_excel(OCTO_dumbi_FILE, game_state.get('used_srno', []))

                if next_answer:
                    game_state.setdefault('used_srno', []).append(next_srno)
                    game_state['current_word'] = next_answer
                    game_state['current_points'] = next_point

                    # Generate the scrambled word and the masked word
                    scrambled_word = ' '.join(random.sample(next_answer, len(next_answer)))
                    #masked_word = mask_word(next_word)  # Call to the mask_word function

                    pass_button = InlineKeyboardButton(text="Hint 🪸", callback_data="hint")
                    keyboard = InlineKeyboardMarkup([[pass_button]])

                    # Send the response with both the scrambled word and masked word
                    try:
                        await update.message.reply_text(
                            f"@{username} earned {formatted_score} points for: {current_word}\n"
                            f"👻 Round: {game_state['current_round']}/{total_rounds}.\n\n"
                            f"☃️ Hint: {next_definition}\n\n"
                            f"🎖️ Points: {next_point}",
                            
                            reply_markup=keyboard
                        )
                    except telegram.error.BadRequest:
                        # If replying fails, send a normal message
                        await update.message.chat.send_message(
                            f"Correct! @{username} earned {formatted_score} points for:→ {current_word}"
                            f"👻 Round: {game_state['current_round']}/{total_rounds}.\n\n"
                            f"☃️ hint: {next_definition}\n\n"
                            f"🎖️ Points: {next_point}",
                            
                            
                            reply_markup=keyboard
                        )
                else:
                    # If no more words are available, end the game
                    try:
                        await update.message.reply_text("No more words available. The game is over.")
                    except telegram.error.BadRequest:
                        await update.message.chat.send_message("No more words available. The game is over.")
                    await show_game_results(update.message, chat_id)  # Show results
                    del octo_game_state[chat_id]

        else:
            # Game is over, show the results
            await show_game_results(update.message, chat_id)
            del octo_game_state[chat_id]
    else:
        # Incorrect guess
        return None

def mask_word(word: str, min_masked: int = 2) -> str:
    """Mask the word by replacing some letters with underscores."""
    if len(word) <= min_masked:
        return '_' * len(word)  # Return all underscores if the word is too short

    # Calculate the minimum number of letters that should remain unmasked
    min_preserved = (len(word)) // 2  # Half of the word length, rounded up for odd lengths
    # Calculate the maximum number of letters to mask
    max_to_mask = len(word) - min_preserved
    # Ensure we mask at least min_masked letters and not more than max_to_mask
    num_to_mask = random.randint(min_masked, max(max_to_mask, min_masked))

    # Select indices to mask
    indices_to_mask = random.sample(range(len(word)), num_to_mask)

    # Create a list of characters from the word
    masked_word_list = list(word)

    # Replace selected indices with underscores
    for index in indices_to_mask:
        masked_word_list[index] = '_ '  # Use a single underscore without space

    # Join the list back into a string
    return ''.join(masked_word_list)

async def handle_round_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):


    query = update.callback_query
    await query.answer()
    chat_id = query.message.chat.id

    if query.data.isdigit():
        selected_rounds = int(query.data)
        if chat_id in octo_game_state:
            try:
                await query.message.reply_text("A game is already running in this group. hit /cancel to end the game")
            except telegram.error.BadRequest:
                await query.message.chat.send_message("A game is already running in this group. hit /cancel to end the game")
            return

        # Initialize the game state for this chat
        srno, word, points = get_random_word_from_excel(OCTO_EXCEL_FILE, [])

        # Shuffle the word and mask it
        scrambled_word = ' '.join(random.sample(word, len(word)))
        masked_word = mask_word(word)  # Call to the mask_word function

        octo_game_state[chat_id] = {
        'total_rounds': selected_rounds,
        'current_round': 1,
        'current_word': word,
        'current_points': points,
        'used_srno': [srno],
        'current_game_score': 0,  # Initialize score for the current game
    }

        pass_button = InlineKeyboardButton(text="Pass 🐛", callback_data="pass")
        keyboard = InlineKeyboardMarkup([[pass_button]])

        try:
            await query.message.reply_text(
            f"Starting game with {selected_rounds} words.\n"
            f"👻 Round:   1/{selected_rounds}.\n"
            f"🎖️ Points:  {points}\n"
            f"📚 Letters:  {scrambled_word}\n"
            f"🎲 Guess:  {masked_word}",
            reply_markup=keyboard
            )
        except telegram.error.BadRequest:
            await query.message.chat.send_message(
            f"Starting game with {selected_rounds} words.\n"
            f"👻 Round:   1/{selected_rounds}.\n"
            f"🎖️ Points:  {points}\n"
            f"📚 Letters:  {scrambled_word}\n"
            f"🎲 Guess:  {masked_word}",
            reply_markup=keyboard
            )
    elif query.data == 'pass':
        # Handle the pass action
        await handle_pass_action(query, chat_id)
    elif query.data == 'hint':
        # Handle the pass action
        await handle_hint_action(query, chat_id)
    elif query.data == 'another_hint':
        # Handle the pass action
        await handle_second_hint_action(query, chat_id)

async def handle_pass_action(query, chat_id):
    print("enter in print 1")
    user_id = query.from_user.id
    username = query.from_user.username or query.from_user.first_name

    if chat_id not in octo_game_state:
        print("enter in print 2")
        await query.message.reply_text("No active game in this group")
        return

    game_state = octo_game_state[chat_id]
    current_word = game_state['current_word']

    # Process passing the current word
    game_state['current_round'] += 1
    print("enter in print 3",dumbai)
    if game_state['current_round'] <= game_state['total_rounds']:

        if dumbai =="dumba":
            print("enter in print 4",dumbai)
            next_srno, next_word, next_points = get_random_word_from_excel(OCTO_EXCEL_FILE, game_state.get('used_srno', []))

            if next_word:
                game_state.setdefault('used_srno', []).append(next_srno)
                game_state['current_word'] = next_word
                game_state['current_points'] = next_points

                scrambled_word = ' '.join(random.sample(next_word, len(next_word)))
                masked_word = mask_word(next_word)
                pass_button = InlineKeyboardButton(text="Pass 🐛", callback_data="pass")
                keyboard = InlineKeyboardMarkup([[pass_button]])

                await query.message.reply_text(
                    f" @{username} passed the word: {current_word}\n"
                    f"👻 Round: {game_state['current_round']}/{game_state['total_rounds']}.\n"
                    f"🎖️ Points: {next_points}\n"
                    f"📚 Letters: {scrambled_word}\n"
                    f"🎲 Guess: {masked_word}\n",
                    reply_markup=keyboard
                )
            else:
                await query.message.reply_text("No more words available The game is over")
                await show_game_results(query.message, chat_id)
                del octo_game_state[chat_id]
        elif dumbai =="dumbii":
            print("enter in print 5",dumbai)
            next_srno,next_definition, next_answer, next_point = dumbi_get_random_word_from_excel(OCTO_dumbi_FILE, game_state.get('used_srno', []))

            if next_answer:
                game_state.setdefault('used_srno', []).append(next_srno)
                game_state['current_word'] = next_answer
                game_state['current_points'] = next_point

                scrambled_word = ' '.join(random.sample(next_answer, len(next_answer)))
                #masked_word = mask_word(next_word)
                pass_button = InlineKeyboardButton(text="Hint 🪸", callback_data="hint")
                keyboard = InlineKeyboardMarkup([[pass_button]])

                await query.message.reply_text(
                    f" @{username} passed the word: {current_word}\n"
                    f"👻 Round: {game_state['current_round']}/{game_state['total_rounds']}.\n\n"
                    f"☃️ Hint: {next_definition}\n\n"
                    f"🎖️ Points: {next_point}",
                    
                    
                    reply_markup=keyboard
                )
            else:
                await query.message.reply_text("No more words available The game is over")
                await show_game_results(query.message, chat_id)
                del octo_game_state[chat_id]

    else:
        await show_game_results(query.message, chat_id)
        del octo_game_state[chat_id]

def escape_markdown(text):
    """Escape special characters in the text for MarkdownV2."""
    if isinstance(text, str):  # Ensure that we're working with a string
        return text.replace('.', '\\.').replace('_', '\\_').replace('*', '\\*').replace('[', '\\[') \
                   .replace(']', '\\]').replace('(', '\\(').replace(')', '\\)').replace('~', '\\~') \
                   .replace('`', '\\`').replace('>', '\\>').replace('#', '\\#').replace('+', '\\+') \
                   .replace('-', '\\-').replace('=', '\\=').replace('|', '\\|').replace('{', '\\{') \
                   .replace('}', '\\}').replace('!', '\\!')
    return str(text)  # Convert non-string types to string

async def show_game_results(message, chat_id):
    if chat_id not in octo_game_state:
        try:
            await message.reply_text("No game in progress")
        except telegram.error.BadRequest:
            await message.chat.send_message("No game in progress")
        return

    game_state = octo_game_state[chat_id]
    players = game_state.get('players', {})

    result_message = "*Game Over*\nScores:\n"

    # Create a sorted list of players based on their current game score in descending order
    sorted_players = sorted(players.items(), key=lambda item: float(item[1]['current_game_score']), reverse=True)

    # Iterate over sorted players and their scores
    for user_id, player_data in sorted_players:
        player_score = player_data['current_game_score']

        # Only show results for players with a score of 1 or more
        if player_score >= 1:
            formatted_score = f"{player_score:.2f}"
            username = escape_markdown(player_data.get('username', 'Unknown User'))  # Handle missing username
            result_message += f"@{username} Score: {escape_markdown(str(formatted_score))} points\n"  # Escape score

    if result_message == "*Game Over*\nScores:\n":
        result_message = "No players with a score of 1 or more"

    try:
        await message.reply_text(result_message, parse_mode='MarkdownV2')
    except telegram.error.BadRequest:
        await message.chat.send_message(result_message, parse_mode='MarkdownV2')




#-------------------------------------------------Start Dumbi     ---------------------------------------------------------------
async def start_dumbii_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.message.chat.id
    global dumbai
    dumbai = "dumbii"
    # Check if the chat_id (group ID) is in the allowed list
    '''
    if chat_id not in ALLOWED_GROUP_IDS:
        try:
            await update.message.reply_text("Due to the free service, you are not allowed to start a game in this group. Play there https://t.me/+yVFKtplWZUA0Yzhl or contact @O000000000O00000000O")
        except telegram.error.BadRequest:
            await update.message.chat.send_message("Due to the free service, you are not allowed to start a game in this group. Play there https://t.me/+yVFKtplWZUA0Yzhl or contact @O000000000O00000000O")
        return
    '''

    if chat_id in octo_game_state:
        try:
            await update.message.reply_text("A game is already running in this group. hit /cancel to end the game")
        except telegram.error.BadRequest:
            await update.message.chat.send_message("A game is already running in this group. hit /cancel to end the game")
        return

    # Present difficulty selection buttons: Easy or Hard
    difficulty_keyboard = [
        [InlineKeyboardButton("Easy Dumbii ❤️‍🔥", callback_data='dumbi_difficulty_easy')],
        [InlineKeyboardButton("Hard Dumbii 💫", callback_data='dumbi_difficulty_hard')],
    ]
    reply_markup = InlineKeyboardMarkup(difficulty_keyboard)

    try:
        await update.message.reply_text('Select the difficulty:', reply_markup=reply_markup)
    except telegram.error.BadRequest:
        await update.message.chat.send_message('Select the difficulty:', reply_markup=reply_markup)
async def handle_hint_action(query, chat_id):
    print("enter in print 1")
    user_id = query.from_user.id
    username = query.from_user.username or query.from_user.first_name

    if chat_id not in octo_game_state:
        print("enter in print 2")
        await query.message.reply_text("No active game in this group")
        return

    game_state = octo_game_state[chat_id]
    current_word = game_state['current_word']
    scrambled_word = ' '.join(random.sample(current_word, len(current_word)))
    #masked_word = mask_word(next_word)
    pass_button = InlineKeyboardButton(text="Another Hint 😤", callback_data="another_hint")
    keyboard = InlineKeyboardMarkup([[pass_button]])

    await query.message.reply_text(
        f" @{username} Asked the hint:\n"
        f"👾 Shuffled letters: {scrambled_word}\n",  
        reply_markup=keyboard
        )
 
async def handle_second_hint_action(query, chat_id):
    print("enter in print 1")
    user_id = query.from_user.id
    username = query.from_user.username or query.from_user.first_name

    if chat_id not in octo_game_state:
        print("enter in print 2")
        await query.message.reply_text("No active game in this group")
        return

    game_state = octo_game_state[chat_id]
    current_word = game_state['current_word']
    scrambled_word = ' '.join(random.sample(current_word, len(current_word)))
    masked_word = mask_word(current_word)
    pass_button = InlineKeyboardButton(text="pass 🌵", callback_data="pass")
    keyboard = InlineKeyboardMarkup([[pass_button]])

    await query.message.reply_text(
        f" @{username} Asked the hint:\n"
        f"🐖 Masked Word: {masked_word}\n"
        f"👾 Shuffled letters: {scrambled_word}\n",  
        reply_markup=keyboard
        )

async def dumbi_difficulty_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
   
    await query.answer()
    chat_id = query.message.chat.id
    if chat_id in octo_game_state:
        try:
            await query.message.reply_text("A game is already running in this group hit /cancel to end the game")
        except telegram.error.BadRequest:
            await query.message.chat.send_message("A game is already running in this group hti /cancel to end the game")
        return
    # Determine difficulty based on callback data
    global OCTO_dumbi_FILE
    difficulty_message = ''
    if query.data == 'dumbi_difficulty_easy':
        excel_files = ['easypuzzl1.xlsx', 'easypuzzl2.xlsx']
    
    # Randomly select a file from the list
        OCTO_dumbi_FILE = random.choice(excel_files)
        difficulty_message = "Easy mode selected"
    elif query.data == 'dumbi_difficulty_hard':
        OCTO_dumbi_FILE = 'puzzl.xlsx'  # Use the hard file
        difficulty_message = "Hard mode selected"

    # Ask how many words to play with
    keyboard = [
        [InlineKeyboardButton("2️⃣ 5️⃣", callback_data='dum_25')],
        [InlineKeyboardButton("1️⃣ 0️⃣ 0️⃣", callback_data='dum_100')],
        [InlineKeyboardButton("2️⃣ 5️⃣ 0️⃣", callback_data='dum_250')],
        [InlineKeyboardButton("5️⃣ 0️⃣ 0️⃣", callback_data='dum_500')],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    # Send a confirmation message along with the next prompt
    try:
        await query.message.reply_text(f"{difficulty_message} How many words do you want", reply_markup=reply_markup)
    except telegram.error.BadRequest:
        await query.message.chat.send_message(f"{difficulty_message} How many words do you want", reply_markup=reply_markup)
def dumbi_get_random_word_from_excel(file_path: str, used_srno: list):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Collect words and points from the Excel file
        words_data = []
        for row in range(2, sheet.max_row + 1):  # Start from the second row to skip headers
            srno = sheet.cell(row=row, column=1).value  # 'srno' is in the first column
            if srno in used_srno:  # Skip words that have already been used
                continue
            definition = sheet.cell(row=row, column=2).value
            answer = sheet.cell(row=row, column=3).value  # Assuming word is in column 2
            point = sheet.cell(row=row, column=4).value  # Assuming points are in column 3

            # Append tuple of srno, word, and points
            words_data.append((srno, definition,answer, point))

        # Choose a random word from the list of unused words
        if words_data:
            srno, definition,answer, point = random.choice(words_data)

            return srno, definition,answer, point
        else:
            return None, None, None

    except FileNotFoundError:
        return None, None, None

def escape_markdown_v2(text: str) -> str:
    return re.sub(r'([_\*\[\]\(\)~`>#+\-=|{}.!])', r'\\\1', text)
    
async def handle_dumbi_round_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):

    query = update.callback_query
    await query.answer()
    chat_id = query.message.chat.id
    round =0
    if query.data == "dum_25":
        round = 25
    elif query.data == "dum_100":
        round = 100
    elif query.data == "dum_250":
        round = 250
    elif query.data == "dum_500":
        round = 500
    if round and round !=0:
        selected_rounds = round
        if chat_id in octo_game_state:
            try:
                await query.message.reply_text("A game is already running in this group. hit /cancel to end the game")
            except telegram.error.BadRequest:
                await query.message.chat.send_message("A game is already running in this group. hit /cancel to end the game")
            return

        # Initialize the game state for this chat
        srno, definition,answer, point = dumbi_get_random_word_from_excel(OCTO_dumbi_FILE, [])

        # Shuffle the word and mask it
        scrambled_word = ' '.join(random.sample(answer, len(answer)))
        #masked_word = mask_word(answer)  # Call to the mask_word function

        octo_game_state[chat_id] = {
        'total_rounds': selected_rounds,
        'current_round': 1,
        'current_word': answer,
        'current_points': point,
        'used_srno': [srno],
        'current_game_score': 0,  # Initialize score for the current game
    }

        pass_button = InlineKeyboardButton(text="Hint 🪸 ", callback_data="hint")
        keyboard = InlineKeyboardMarkup([[pass_button]])

        try:
            await query.message.reply_text(
            f"Starting game with {selected_rounds} Rounds.\n"
            f"👻 Round:   1/{selected_rounds}.\n\n"
            f"☃️ Hint:  {definition}\n\n"
            f"🎖️ Points:  {point}",
            reply_markup=keyboard
            )
        except telegram.error.BadRequest:
            await query.message.chat.send_message(
            f"Starting game with {selected_rounds} Rounds.\n\n"
            f"👻 Round:   1/{selected_rounds}.\n"
            f"☃️ Hint:  {definition}\n\n"
            f"🎖️ Points:  {point}",
            
            reply_markup=keyboard
            )
    elif query.data == 'pass':
        # Handle the pass action
        await handle_pass_action(query, chat_id)
#-----------------------------------download score table ---------------------------------------------
async def download_scores_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        chat_id = update.message.chat.id
        
        # Check if the chat_id (group ID) is in the allowed list
    
        if chat_id not in ALLOWED_GROUP_IDS:
            try:
                await update.message.reply_text("Due to the free service, you are not allowed to start a game in this group. Play there https://t.me/+yVFKtplWZUA0Yzhl or contact @O000000000O00000000O")
            except telegram.error.BadRequest:
                await update.message.chat.send_message("Due to the free service, you are not allowed to start a game in this group. Play there https://t.me/+yVFKtplWZUA0Yzhl or contact @O000000000O00000000O")
            return

        # Check if the file exists
        if os.path.exists(EXCEL_FILE):
            # Send the file to the user
            with open(EXCEL_FILE, 'rb') as file:
                await context.bot.send_document(chat_id=update.message.chat.id, document=file)
        else:
            # Notify the user that the file does not exist
            await update.message.reply_text("Sorry, the score file is not available.")
    except Exception as e:
        # Handle any errors
        await update.message.reply_text(f"An error occurred: {e}")
#-------------------------------------------------Start Dumbi     ---------------------------------------------------------------
def main():
    # Create the application
    application = Application.builder().token(TOKEN).build()

    # Register handlers
    application.add_handler(CommandHandler('startdumba', start_game_command))
    application.add_handler(CallbackQueryHandler(handle_difficulty_selection, pattern='^difficulty_'))
    application.add_handler(CallbackQueryHandler(handle_round_selection, pattern='^pass'))
    application.add_handler(CallbackQueryHandler(handle_round_selection, pattern='^\d+$'))


    application.add_handler(CommandHandler('startdumbii', start_dumbii_command))
    application.add_handler(CallbackQueryHandler(dumbi_difficulty_selection, pattern='^dumbi_'))
    application.add_handler(CallbackQueryHandler(handle_round_selection, pattern='^hint'))
    application.add_handler(CallbackQueryHandler(handle_round_selection, pattern='^another_hint'))
    application.add_handler(CallbackQueryHandler(handle_dumbi_round_selection, pattern='^dum_'))
    
    
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, process_game_round))
    application.add_handler(CommandHandler('cancel', cancel_game))
    application.add_handler(CommandHandler('showallresults', show_all_results))
    application.add_handler(CommandHandler('myrank', my_rank))
    application.add_handler(CommandHandler('top10dumb', select_top_10_users))

    application.add_handler(CommandHandler('downloadscoreiesp', download_scores_command))
    


    # Start the bot
    application.run_polling()


if __name__ == '__main__':
    main()
